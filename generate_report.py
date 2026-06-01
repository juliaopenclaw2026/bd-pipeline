"""
generate_report.py
Generates SAM_Opportunities_Report_YYYYMMDD.xlsx from the latest syte_opportunities CSV.
CEO-level BD pipeline report with 4 sheets.

Usage (standalone):
    python generate_report.py

Called by run.py with:
    generate_report.main(csv_path, today)
"""
import csv
import re
import os
import glob
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from urllib.parse import quote as url_quote


def get_display_link(row):
    """Return the best available clickable URL for an opportunity row.
    Priority: uiLink (SAM direct) > source_link > SAM search fallback."""
    link = row.get("uiLink", "") or row.get("source_link", "")
    if link and str(link) not in ("", "nan"):
        return str(link)
    sol = str(row.get("solicitationNumber", "")).strip()
    if sol:
        return f"https://sam.gov/search/?keywords={url_quote(sol)}&index=opp"
    return ""


# ── Runtime config (set by main(), overrideable) ──────────────────────────────
TODAY = datetime.now()
INPUT_CSV = None   # resolved at runtime
OUTPUT_FILE = None  # resolved at runtime

# Color palette
C_NAVY       = "1B2A4A"
C_STEEL      = "2E86AB"
C_GOLD       = "F18F01"
C_GREEN      = "44BBA4"
C_RED        = "E94F37"
C_LIGHT_BG   = "F5F7FA"
C_ALT_ROW    = "EBF3FB"
C_WHITE      = "FFFFFF"
C_DARK_TEXT  = "1B2A4A"
C_MID_TEXT   = "4A5568"

NAICS_LABELS = {
    "237110": "237110 — Water & Sewer",
    "237120": "237120 — Oil & Gas Pipeline",
    "237310": "237310 — Highway & Bridge",
    "237990": "237990 — Other Heavy/Civil",
    "213112": "213112 — Support Activities Oil & Gas",
    "237130": "237130 — Power & Communication Line",
    "238210": "238210 — Electrical Contractors",
    "561210": "561210 — Facilities Maintenance",
}

SPECIALTY_NAICS = {"213112", "237120", "237110", "237130", "238210", "561210"}

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color=C_DARK_TEXT, size=10, bold=False, italic=False):
    return Font(name="Calibri", color=color, size=size, bold=bold, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides="bottom"):
    s = Side(style="thin", color="D1D5DB")
    b = Border(
        left=s if "left" in sides else Side(),
        right=s if "right" in sides else Side(),
        top=s if "top" in sides else Side(),
        bottom=s if "bottom" in sides else Side(),
    )
    return b

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def parse_rom(rom_text):
    if not rom_text or str(rom_text).strip() == "":
        return None
    suffix_map = {"thousand": 1e3, "k": 1e3, "million": 1e6, "m": 1e6,
                  "billion": 1e9, "b": 1e9}
    matches = re.findall(
        r'\$([\d,]+(?:\.\d+)?)\s*(million|billion|thousand|M|B|K)?',
        str(rom_text), re.IGNORECASE)
    amounts = []
    for num, suf in matches:
        try:
            v = float(num.replace(",", ""))
            if suf:
                v *= suffix_map.get(suf.lower(), 1)
            amounts.append(v)
        except ValueError:
            pass
    return sum(amounts) / len(amounts) if amounts else None

def fmt_rom(value):
    if value is None:
        return "—"
    if value >= 1e9:
        return f"${value/1e9:.1f}B"
    if value >= 1e6:
        return f"${value/1e6:.1f}M"
    if value >= 1e3:
        return f"${value/1e3:.0f}K"
    return f"${value:,.0f}"

def parse_deadline(s):
    if not s or str(s).strip() in ("", "nan"):
        return None
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d")
    except ValueError:
        return None

def days_left(deadline_dt):
    if deadline_dt is None:
        return None
    return (deadline_dt - TODAY).days

def urgency_label(dl):
    if dl is None:
        return ("Pre-Sol", "718096", "FFFFFF")
    if dl <= 14:
        return ("🔴 Urgent", "E94F37", "FFFFFF")
    if dl <= 30:
        return ("🟡 Soon", "F18F01", "FFFFFF")
    return ("⬜ Watch", "718096", "FFFFFF")

def is_sdvosb(row):
    desc = str(row.get("typeOfSetAsideDescription", "")).lower()
    return "veteran" in desc or "sdvosb" in desc

def shorten(text, n=60):
    return text[:n] + "…" if len(text) > n else text

def dept_label(row):
    fp = str(row.get("fullParentPathName", "") or "")
    if fp and fp != "nan":
        return fp.split(".")[0].strip().title()
    d = str(row.get("department", "") or "")
    if d and d != "nan":
        return d.split(".")[0].strip().title()
    return "—"

# ── Load data ─────────────────────────────────────────────────────────────────
def load_data(csv_path=None):
    if csv_path is None:
        csv_path = INPUT_CSV
    rows = []
    with open(csv_path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            row["_deadline_dt"] = parse_deadline(row.get("responseDeadLine", ""))
            row["_days_left"] = days_left(row["_deadline_dt"])
            row["_rom_value"] = parse_rom(row.get("ROM", ""))
            row["_is_sdvosb"] = is_sdvosb(row)
            # Parse LLM score (may be absent on first run without API key)
            raw_score = row.get("LLM_Score", "")
            row["_llm_score"] = int(raw_score) if str(raw_score).isdigit() else None
            # Exclude rows where deadline has already passed
            dl = days_left(row["_deadline_dt"])
            if row["_deadline_dt"] is not None and dl < 0:
                continue
            rows.append(row)
    return rows

# ── Sheet 1: Executive Dashboard ──────────────────────────────────────────────
def build_dashboard(ws, rows):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_NAVY

    # Column widths
    widths = {"A": 2, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 2}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 44
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 12

    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "SYTE Corp — BD Pipeline Report"
    c.font = Font(name="Calibri", size=22, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = align("center", "center")

    ws.merge_cells("B3:G3")
    c = ws["B3"]
    c.value = f"Prepared by BD Team  ·  Report Date: {TODAY.strftime('%B %d, %Y')}  ·  Data Source: SAM.gov"
    c.font = Font(name="Calibri", size=10, color="A8C4E0", italic=True)
    c.fill = fill(C_STEEL)
    c.alignment = align("center", "center")

    # ── KPI Cards (row 5–8) ───────────────────────────────────────────────────
    ws.row_dimensions[5].height = 8
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 32
    ws.row_dimensions[8].height = 18
    ws.row_dimensions[9].height = 12

    solicitations = [r for r in rows if r["type"] == "Solicitation"]
    presols = [r for r in rows if r["type"] == "Presolicitation"]
    sdvosb_count = sum(1 for r in rows if r["_is_sdvosb"])
    urgent = sum(1 for r in solicitations if r["_days_left"] is not None and r["_days_left"] <= 14)
    soon = sum(1 for r in solicitations if r["_days_left"] is not None and 15 <= r["_days_left"] <= 30)
    total_rom = sum(r["_rom_value"] for r in rows if r["_rom_value"])

    kpis = [
        ("Total Opportunities", str(len(rows)), C_STEEL),
        ("Pipeline Value (ROM)", fmt_rom(total_rom), C_GREEN),
        ("🔴 Urgent (≤14d)", str(urgent), C_RED),
        ("🟡 Soon (15–30d)", str(soon), C_GOLD),
        ("SDVOSB Set-Aside", str(sdvosb_count), "8B6914"),
        ("Active Solicitations", str(len(solicitations)), C_STEEL),
    ]
    cols = ["B", "C", "D", "E", "F", "G"]

    for col, (label, value, accent) in zip(cols, kpis):
        # Label row
        lc = ws[f"{col}6"]
        lc.value = label
        lc.font = Font(name="Calibri", size=8, bold=True, color="718096")
        lc.fill = fill(C_LIGHT_BG)
        lc.alignment = align("center", "center")

        # Value row
        vc = ws[f"{col}7"]
        vc.value = value
        vc.font = Font(name="Calibri", size=20, bold=True, color=accent)
        vc.fill = fill(C_LIGHT_BG)
        vc.alignment = align("center", "center")

        # Bottom accent bar
        bc = ws[f"{col}8"]
        bc.fill = fill(accent)

    # ── Top 5 Urgent Actions (rows 10–17) ────────────────────────────────────
    ws.row_dimensions[10].height = 24
    ws.merge_cells("B10:G10")
    hdr = ws["B10"]
    hdr.value = "⚡  IMMEDIATE ACTION REQUIRED — Upcoming Deadlines"
    hdr.font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
    hdr.fill = fill(C_RED)
    hdr.alignment = align("left", "center")

    # Sub-header
    ws.row_dimensions[11].height = 18
    for col, label in zip(["B", "C", "D", "E", "F", "G"],
                           ["Solicitation #", "Title", "NAICS", "ROM", "Deadline", "Days Left"]):
        c = ws[f"{col}11"]
        c.value = label
        c.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
        c.fill = fill(C_NAVY)
        c.alignment = align("center", "center")

    urgent_rows = sorted(
        [r for r in solicitations if r["_days_left"] is not None and r["_days_left"] >= 0],
        key=lambda r: r["_days_left"]
    )[:5]

    for i, row in enumerate(urgent_rows):
        r = 12 + i
        ws.row_dimensions[r].height = 20
        bg = C_WHITE if i % 2 == 0 else C_ALT_ROW
        sdv_bg = "FFFBEC" if row["_is_sdvosb"] else bg

        data = [
            ("B", row.get("solicitationNumber", "—")),
            ("C", shorten(row.get("title", "—"), 52)),
            ("D", str(row.get("naicsCode", "—"))),
            ("E", fmt_rom(row["_rom_value"])),
            ("F", row["_deadline_dt"].strftime("%b %d, %Y") if row["_deadline_dt"] else "—"),
            ("G", f"{row['_days_left']}d"),
        ]
        for col, val in data:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Calibri", size=9, color=C_DARK_TEXT,
                          bold=(col == "G"))
            c.fill = fill(sdv_bg)
            c.alignment = align("left" if col in ("B", "C") else "center", "center")

        # Color days left cell
        dl = row["_days_left"]
        gc = ws[f"G{r}"]
        if dl <= 7:
            gc.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
            gc.fill = fill(C_RED)
        elif dl <= 14:
            gc.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
            gc.fill = fill("E07B3A")

        # SDVOSB badge
        if row["_is_sdvosb"]:
            c = ws[f"B{r}"]
            c.font = Font(name="Calibri", size=9, bold=True, color="8B6914")

    # ── Breakdown tables (rows 19+) ───────────────────────────────────────────
    row_start = 18 + len(urgent_rows) + 2

    def section_header(ws, row, label, cols="B:G"):
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"B{row}:G{row}")
        c = ws[f"B{row}"]
        c.value = label
        c.font = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
        c.fill = fill(C_STEEL)
        c.alignment = align("left", "center")

    def table_row(ws, row, label, value, i):
        ws.row_dimensions[row].height = 18
        bg = C_LIGHT_BG if i % 2 == 0 else C_WHITE
        lc = ws[f"B{row}"]
        lc.value = label
        lc.font = Font(name="Calibri", size=9, color=C_DARK_TEXT)
        lc.fill = fill(bg)
        lc.alignment = align("left", "center")
        ws.merge_cells(f"B{row}:E{row}")

        vc = ws[f"F{row}"]
        vc.value = value
        vc.font = Font(name="Calibri", size=11, bold=True, color=C_STEEL)
        vc.fill = fill(bg)
        vc.alignment = align("center", "center")
        ws.merge_cells(f"F{row}:G{row}")

    # By Stage
    section_header(ws, row_start, "📋  BY STAGE")
    row_start += 1
    stage_data = [
        ("Solicitation (Active Bid)", len(solicitations)),
        ("Presolicitation (Watch)", len(presols)),
        ("Special Notice", sum(1 for r in rows if r["type"] == "Special Notice")),
    ]
    for i, (label, val) in enumerate(stage_data):
        table_row(ws, row_start + i, label, str(val), i)
    row_start += len(stage_data) + 1

    # By Set-Aside
    section_header(ws, row_start, "🎖  BY SET-ASIDE TYPE")
    row_start += 1
    from collections import Counter
    sa_counts = Counter(
        "SDVOSB" if r["_is_sdvosb"] else
        ("SBA (Small Business)" if "small business" in str(r.get("typeOfSetAsideDescription","")).lower() else "Other")
        for r in rows
    )
    for i, (label, val) in enumerate([
        ("SDVOSB (Prime Target)", sa_counts["SDVOSB"]),
        ("SBA — Total Small Business", sa_counts["SBA (Small Business)"]),
        ("Other", sa_counts["Other"]),
    ]):
        table_row(ws, row_start + i, label, str(val), i)
    row_start += 4

    # By NAICS
    section_header(ws, row_start, "🏗  BY NAICS CODE")
    row_start += 1
    naics_counts = Counter(str(r.get("naicsCode", "")) for r in rows)
    for i, code in enumerate(["237110", "237310", "237990", "237120"]):
        table_row(ws, row_start + i, NAICS_LABELS.get(code, code), str(naics_counts.get(code, 0)), i)
    row_start += len(NAICS_LABELS) + 1

    # Key Insights
    section_header(ws, row_start, "💡  KEY INSIGHTS")
    row_start += 1

    insights = []
    insights.append(f"→  {urgent} active Solicitation(s) due within 14 days — immediate bid/no-bid decisions required.")
    sdvosb_sol = [r for r in solicitations if r["_is_sdvosb"]]
    if sdvosb_sol:
        insights.append(f"→  {len(sdvosb_sol)} SDVOSB Solicitation(s) identified — strategic priority for SYTE Corp.")
    top_rom = sorted([r for r in rows if r["_rom_value"]], key=lambda r: -r["_rom_value"])
    if top_rom:
        tr = top_rom[0]
        insights.append(f"→  Highest-value opportunity: {shorten(tr['title'], 50)} (est. {fmt_rom(tr['_rom_value'])}).")

    for i, insight in enumerate(insights):
        ws.row_dimensions[row_start + i].height = 20
        ws.merge_cells(f"B{row_start+i}:G{row_start+i}")
        c = ws[f"B{row_start+i}"]
        c.value = insight
        c.font = Font(name="Calibri", size=9, color=C_DARK_TEXT, italic=(i > 0))
        c.fill = fill("FFFBEC" if i == 0 else (C_LIGHT_BG if i % 2 == 0 else C_WHITE))
        c.alignment = align("left", "center", wrap=True)


# ── Sheet 2: Action Required ──────────────────────────────────────────────────
def build_action_required(ws, rows):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_RED

    solicitations = [r for r in rows if r["type"] == "Solicitation"]
    # Primary: urgency bucket (~weekly); secondary: AI score desc; tertiary: exact deadline asc
    solicitations.sort(key=lambda r: (
        r["_days_left"] is None,
        (r["_days_left"] or 9999) // 7,
        -(r["_llm_score"] or 0),
        r["_days_left"] or 9999,
    ))

    widths = {"A": 2, "B": 5, "C": 18, "D": 44, "E": 20, "F": 12,
              "G": 20, "H": 14, "I": 14, "J": 10, "K": 12, "L": 2}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    # Title
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:N2")
    c = ws["B2"]
    c.value = "🔴  Action Required — Active Solicitations"
    c.font = Font(name="Calibri", size=16, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = align("left", "center")

    ws.row_dimensions[3].height = 18
    ws.merge_cells("B3:N3")
    c = ws["B3"]
    c.value = (f"All open Solicitations sorted by urgency  ·  "
               f"{len(solicitations)} records  ·  As of {TODAY.strftime('%B %d, %Y')}  ·  "
               f"★ Gold rows = SDVOSB set-aside")
    c.font = Font(name="Calibri", size=9, color=C_WHITE, italic=True)
    c.fill = fill(C_STEEL)
    c.alignment = align("left", "center")

    ws.row_dimensions[4].height = 8

    # Headers
    headers = ["#", "Sol #", "Title", "Location", "NAICS", "Set-Aside",
               "ROM", "Deadline", "Days Left", "Urgency", "AI Score", "AI Reason", "SAM Link"]
    header_cols = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
    widths2 = {"A":2,"B":5,"C":18,"D":44,"E":20,"F":12,"G":18,"H":12,"I":14,"J":10,"K":14,"L":7,"M":35,"N":22,"O":2}
    for col, w in widths2.items():
        set_col_width(ws, col, w)

    ws.row_dimensions[5].height = 22
    for col, label in zip(header_cols, headers):
        c = ws[f"{col}5"]
        c.value = label
        c.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
        c.fill = fill(C_NAVY)
        c.alignment = align("center", "center")

    ws.freeze_panes = "B6"

    for i, row in enumerate(solicitations):
        r = 6 + i
        ws.row_dimensions[r].height = 32
        sdvosb = row["_is_sdvosb"]
        bg = ("FFFBEC" if sdvosb else (C_WHITE if i % 2 == 0 else C_ALT_ROW))

        dl = row["_days_left"]
        urgency, urg_bg, urg_fg = urgency_label(dl)
        deadline_dt = row["_deadline_dt"]

        cells = [
            ("B", str(i + 1)),
            ("C", row.get("solicitationNumber", "—")),
            ("D", row.get("title", "—")),
            ("E", row.get("location", "—")),
            ("F", str(row.get("naicsCode", "—"))),
            ("G", "SDVOSB ★" if sdvosb else "SBA"),
            ("H", fmt_rom(row["_rom_value"])),
            ("I", deadline_dt.strftime("%b %d, %Y") if deadline_dt else "—"),
            ("J", str(dl) + "d" if dl is not None else "—"),
        ]
        for col, val in cells:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Calibri", size=9,
                          color=("8B6914" if (sdvosb and col == "G") else C_DARK_TEXT),
                          bold=(col in ("B", "G") and sdvosb))
            c.fill = fill(bg)
            c.alignment = align("left" if col == "D" else "center", "center", wrap=(col == "D"))

        # Urgency badge (K)
        uc = ws[f"K{r}"]
        uc.value = urgency
        uc.font = Font(name="Calibri", size=9, bold=True, color=urg_fg)
        uc.fill = fill(urg_bg)
        uc.alignment = align("center", "center")

        # AI Score (L)
        ai_score = row.get("_llm_score")
        lc = ws[f"L{r}"]
        if ai_score is not None:
            lc.value = ai_score
            score_color = (C_GREEN if ai_score >= 7 else (C_GOLD if ai_score >= 4 else C_RED))
            lc.font = Font(name="Calibri", size=9, bold=True, color=score_color)
        else:
            lc.value = "—"
            lc.font = Font(name="Calibri", size=9, color=C_DARK_TEXT)
        lc.fill = fill(bg)
        lc.alignment = align("center", "center")

        # AI Reason (M)
        mc = ws[f"M{r}"]
        mc.value = row.get("LLM_Reason", "—") or "—"
        mc.font = Font(name="Calibri", size=8, color=C_MID_TEXT, italic=True)
        mc.fill = fill(bg)
        mc.alignment = align("left", "center", wrap=True)

        # SAM Link (N)
        link = get_display_link(row)
        nc = ws[f"N{r}"]
        if link:
            nc.value = "SAM.gov ↗"
            nc.hyperlink = link
            nc.font = Font(name="Calibri", size=9, color="0563C1", underline="single")
        else:
            nc.value = "—"
            nc.font = Font(name="Calibri", size=9, color=C_DARK_TEXT)
        nc.fill = fill(bg)
        nc.alignment = align("center", "center")

    ws.auto_filter.ref = f"B5:N{5 + len(solicitations)}"


# ── Sheet 3: Presolicitation Watch List ───────────────────────────────────────
def build_watch_list(ws, rows):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_GOLD

    presols = [r for r in rows if r["type"] in ("Presolicitation", "Special Notice")]
    # SDVOSB first, then AI score desc, then ROM desc
    presols.sort(key=lambda r: (not r["_is_sdvosb"], -(r["_llm_score"] or 0), -(r["_rom_value"] or 0)))

    widths = {"A":2,"B":18,"C":44,"D":20,"E":12,"F":18,"G":14,"H":16,"I":10,"J":7,"K":22,"L":2}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:K2")
    c = ws["B2"]
    c.value = "👀  Presolicitation Watch List — Pipeline Intelligence"
    c.font = Font(name="Calibri", size=16, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = align("left", "center")

    ws.row_dimensions[3].height = 18
    ws.merge_cells("B3:K3")
    c = ws["B3"]
    sdvosb_count = sum(1 for r in presols if r["_is_sdvosb"])
    c.value = (f"Early-stage opportunities to monitor  ·  "
               f"{len(presols)} records  ({sdvosb_count} SDVOSB)  ·  "
               f"SDVOSB first, then AI score desc")
    c.font = Font(name="Calibri", size=9, color=C_WHITE, italic=True)
    c.fill = fill(C_STEEL)
    c.alignment = align("left", "center")

    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 22

    headers = ["Sol #", "Title", "Location", "NAICS", "Set-Aside", "ROM Est.", "Est. Deadline", "Days Left", "Type", "AI Score"]
    for col, label in zip(list("BCDEFGHIJK"), headers):
        c = ws[f"{col}5"]
        c.value = label
        c.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
        c.fill = fill(C_NAVY)
        c.alignment = align("center", "center")

    ws.freeze_panes = "B6"

    for i, row in enumerate(presols):
        r = 6 + i
        ws.row_dimensions[r].height = 32
        sdvosb = row["_is_sdvosb"]
        bg = "FFFBEC" if sdvosb else (C_WHITE if i % 2 == 0 else C_ALT_ROW)

        dl = row["_days_left"]
        deadline_dt = row["_deadline_dt"]

        cells = [
            ("B", row.get("solicitationNumber", "—")),
            ("C", row.get("title", "—")),
            ("D", row.get("location", "—")),
            ("E", str(row.get("naicsCode", "—"))),
            ("F", "SDVOSB ★" if sdvosb else "SBA"),
            ("G", fmt_rom(row["_rom_value"])),
            ("H", deadline_dt.strftime("%b %d, %Y") if deadline_dt else "TBD"),
            ("I", (str(dl) + "d") if dl is not None else "TBD"),
            ("J", row.get("type", "—")),
        ]
        for col, val in cells:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Calibri", size=9,
                          color=("8B6914" if (sdvosb and col == "F") else C_DARK_TEXT),
                          bold=(sdvosb and col in ("B", "F")))
            c.fill = fill(bg)
            c.alignment = align("left" if col == "C" else "center", "center", wrap=(col == "C"))

        # AI Score (K)
        ai_score = row.get("_llm_score")
        kc = ws[f"K{r}"]
        if ai_score is not None:
            kc.value = ai_score
            score_color = (C_GREEN if ai_score >= 7 else (C_GOLD if ai_score >= 4 else C_RED))
            kc.font = Font(name="Calibri", size=9, bold=True, color=score_color)
        else:
            kc.value = "—"
            kc.font = Font(name="Calibri", size=9, color=C_DARK_TEXT)
        kc.fill = fill(bg)
        kc.alignment = align("center", "center")

    ws.auto_filter.ref = f"B5:K{5 + len(presols)}"


# ── Sheet 4: Full Pipeline ─────────────────────────────────────────────────────
def build_full_pipeline(ws, rows):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_STEEL

    widths = {"A":2,"B":12,"C":44,"D":20,"E":12,"F":18,"G":14,"H":16,"I":10,"J":14,"K":22,"L":2}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 32
    ws.merge_cells("B2:K2")
    c = ws["B2"]
    c.value = "📋  Full Pipeline — All Opportunities"
    c.font = Font(name="Calibri", size=14, bold=True, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = align("left", "center")

    ws.row_dimensions[3].height = 16
    ws.merge_cells("B3:K3")
    c = ws["B3"]
    c.value = f"Complete data  ·  {len(rows)} records  ·  {TODAY.strftime('%B %d, %Y')}"
    c.font = Font(name="Calibri", size=9, color=C_WHITE, italic=True)
    c.fill = fill(C_STEEL)
    c.alignment = align("left", "center")

    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 20

    headers = ["Sol #", "Title", "Location", "NAICS", "Set-Aside", "ROM", "Deadline", "Days Left", "Type", "SAM Link"]
    for col, label in zip(list("BCDEFGHIJK"), headers):
        c = ws[f"{col}5"]
        c.value = label
        c.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
        c.fill = fill(C_NAVY)
        c.alignment = align("center", "center")

    ws.freeze_panes = "B6"

    # Sort: Solicitation first, then Presolicitation, then others; within each group by days_left
    sorted_rows = sorted(rows, key=lambda r: (
        {"Solicitation": 0, "Presolicitation": 1, "Special Notice": 2}.get(r["type"], 3),
        r["_days_left"] is None,
        r["_days_left"] or 9999
    ))

    for i, row in enumerate(sorted_rows):
        r = 6 + i
        ws.row_dimensions[r].height = 28
        sdvosb = row["_is_sdvosb"]
        bg = "FFFBEC" if sdvosb else (C_WHITE if i % 2 == 0 else C_ALT_ROW)

        dl = row["_days_left"]
        deadline_dt = row["_deadline_dt"]
        link = get_display_link(row)

        cells = [
            ("B", row.get("solicitationNumber", "—")),
            ("C", row.get("title", "—")),
            ("D", row.get("location", "—")),
            ("E", str(row.get("naicsCode", "—"))),
            ("F", "SDVOSB ★" if sdvosb else "SBA"),
            ("G", fmt_rom(row["_rom_value"])),
            ("H", deadline_dt.strftime("%b %d, %Y") if deadline_dt else "—"),
            ("I", (str(dl) + "d") if dl is not None else "—"),
            ("J", row.get("type", "—")),
        ]
        for col, val in cells:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Calibri", size=9, color=C_DARK_TEXT)
            c.fill = fill(bg)
            c.alignment = align("left" if col == "C" else "center", "center", wrap=(col == "C"))

        # SAM Link
        kc = ws[f"K{r}"]
        if link:
            kc.value = "SAM.gov ↗"
            kc.hyperlink = link
            kc.font = Font(name="Calibri", size=9, color="0563C1", underline="single")
        else:
            kc.value = "—"
            kc.font = Font(name="Calibri", size=9, color=C_DARK_TEXT)
        kc.fill = fill(bg)
        kc.alignment = align("center", "center")

    ws.auto_filter.ref = f"B5:K{5 + len(rows)}"


# ── Sheet 5: Specialty Pipeline ───────────────────────────────────────────────
def build_specialty_pipeline(ws, rows):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8B5CF6"  # purple

    # Filter to specialty NAICS only
    spec_rows = [r for r in rows if str(r.get("naicsCode", "")) in SPECIALTY_NAICS]
    spec_rows.sort(key=lambda r: (
        not r["_is_sdvosb"],
        {"Solicitation": 0, "Presolicitation": 1}.get(r["type"], 2),
        r["_days_left"] is None,
        r["_days_left"] or 9999,
    ))

    widths = {"A":2,"B":22,"C":44,"D":20,"E":12,"F":18,"G":14,"H":16,"I":10,"J":14,"K":22,"L":2}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 32
    ws.merge_cells("B2:K2")
    c = ws["B2"]
    c.value = "🔧  Specialty Pipeline — Oil, Gas, Utilities & Facilities"
    c.font = Font(name="Calibri", size=14, bold=True, color=C_WHITE)
    c.fill = fill("8B5CF6")
    c.alignment = align("left", "center")

    ws.row_dimensions[3].height = 16
    ws.merge_cells("B3:K3")
    c = ws["B3"]
    sdvosb_count = sum(1 for r in spec_rows if r["_is_sdvosb"])
    c.value = (f"NAICS: 213112 · 237110 · 237120 · 237130 · 238210 · 561210  ·  "
               f"{len(spec_rows)} records  ({sdvosb_count} SDVOSB)  ·  {TODAY.strftime('%B %d, %Y')}")
    c.font = Font(name="Calibri", size=9, color=C_WHITE, italic=True)
    c.fill = fill("7C3AED")
    c.alignment = align("left", "center")

    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 20

    headers = ["NAICS", "Title", "Location", "Set-Aside", "ROM", "Deadline", "Days Left", "Type", "Urgency", "SAM Link"]
    for col, label in zip(list("BCDEFGHIJK"), headers):
        c = ws[f"{col}5"]
        c.value = label
        c.font = Font(name="Calibri", size=9, bold=True, color=C_WHITE)
        c.fill = fill("8B5CF6")
        c.alignment = align("center", "center")

    ws.freeze_panes = "B6"

    if not spec_rows:
        ws.merge_cells("B6:K6")
        c = ws["B6"]
        c.value = "No opportunities found for these NAICS codes in the current dataset."
        c.font = Font(name="Calibri", size=10, color=C_MID_TEXT, italic=True)
        c.alignment = align("center", "center")
        return

    for i, row in enumerate(spec_rows):
        r = 6 + i
        ws.row_dimensions[r].height = 28
        sdvosb = row["_is_sdvosb"]
        bg = "F5F3FF" if sdvosb else (C_WHITE if i % 2 == 0 else "F5F3FF")
        bg = "EDE9FE" if sdvosb else (C_WHITE if i % 2 == 0 else "F5F3FF")

        dl = row["_days_left"]
        deadline_dt = row["_deadline_dt"]
        urgency, urg_bg, urg_fg = urgency_label(dl)
        link = get_display_link(row)

        naics_code = str(row.get("naicsCode", "—"))
        naics_label = NAICS_LABELS.get(naics_code, naics_code)

        cells = [
            ("B", naics_label),
            ("C", row.get("title", "—")),
            ("D", row.get("location", "—")),
            ("E", "SDVOSB ★" if sdvosb else "SBA"),
            ("F", fmt_rom(row["_rom_value"])),
            ("G", deadline_dt.strftime("%b %d, %Y") if deadline_dt else "—"),
            ("H", (str(dl) + "d") if dl is not None else "—"),
            ("I", row.get("type", "—")),
        ]
        for col, val in cells:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Calibri", size=9,
                          color=("8B5CF6" if (sdvosb and col == "E") else C_DARK_TEXT),
                          bold=(sdvosb and col == "E"))
            c.fill = fill(bg)
            c.alignment = align("left" if col in ("B", "C") else "center", "center", wrap=(col == "C"))

        # Urgency badge (J)
        jc = ws[f"J{r}"]
        jc.value = urgency
        jc.font = Font(name="Calibri", size=9, bold=True, color=urg_fg)
        jc.fill = fill(urg_bg)
        jc.alignment = align("center", "center")

        # SAM Link (K)
        kc = ws[f"K{r}"]
        if link:
            kc.value = "SAM.gov ↗"
            kc.hyperlink = link
            kc.font = Font(name="Calibri", size=9, color="0563C1", underline="single")
        else:
            kc.value = "—"
            kc.font = Font(name="Calibri", size=9, color=C_DARK_TEXT)
        kc.fill = fill(bg)
        kc.alignment = align("center", "center")

    ws.auto_filter.ref = f"B5:K{5 + len(spec_rows)}"


# ── Main ──────────────────────────────────────────────────────────────────────
def main(csv_path=None, today=None):
    global TODAY
    base = os.path.dirname(os.path.abspath(__file__))

    # Resolve TODAY
    if today is not None:
        TODAY = today
    else:
        TODAY = datetime.now()

    # Resolve input CSV
    if csv_path is None:
        pattern = os.path.join(base, "syte_opportunities_????????.csv")
        candidates = sorted(glob.glob(pattern))
        csv_path = candidates[-1] if candidates else os.path.join(base, "syte_opportunities.csv")

    # Resolve output filename (AM/PM suffix so morning & afternoon runs don't overwrite)
    slot = "AM" if TODAY.hour < 12 else "PM"
    date_str = f"{TODAY.strftime('%Y%m%d')}_{slot}"
    out_path = os.path.join(base, f"SAM_Opportunities_Report_{date_str}.xlsx")

    print("Loading data...")
    rows = load_data(csv_path)
    print(f"  {len(rows)} rows loaded from {os.path.basename(csv_path)}.")

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "📊 Executive Dashboard"
    print("Building Executive Dashboard...")
    build_dashboard(ws1, rows)

    ws2 = wb.create_sheet("🔴 Action Required")
    print("Building Action Required sheet...")
    build_action_required(ws2, rows)

    ws3 = wb.create_sheet("👀 Watch List")
    print("Building Presolicitation Watch List...")
    build_watch_list(ws3, rows)

    ws4 = wb.create_sheet("📋 Full Pipeline")
    print("Building Full Pipeline...")
    build_full_pipeline(ws4, rows)

    ws5 = wb.create_sheet("🔧 Specialty Pipeline")
    print("Building Specialty Pipeline...")
    build_specialty_pipeline(ws5, rows)

    wb.save(out_path)
    print(f"\n✅ Report saved: {out_path}")


if __name__ == "__main__":
    main()
